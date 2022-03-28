from io import FileIO
from pathlib import Path
from typing import Any, AnyStr
import openpyxl
from pprint import pprint
from prettytable import PrettyTable

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .styles import Colors, FontStyles


# https://openpyxl.readthedocs.io/en/stable/tutorial.html
# https://habr.com/ru/company/otus/blog/331998/


class ExcelReader:

    def __init__(self, filename:str):
        self.filename = str(Path(Path.cwd()).joinpath(filename))
        self.workbook:Workbook
        self.worksheet:Worksheet

    def open_file(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            print('Файл прочитан')
        except Exception as e:
            raise FileNotFoundError(e.args)

    
    def read_sheet(self):
        if not self.workbook:
            raise AttributeError("Файл закрыт")
        self.worksheet = self.workbook.active
        print(f'Рабочий лист {self.worksheet.title} загружен')


    def read_single_cell(self):
        # чтение ячейки
        if not self.worksheet:
            raise AttributeError("Рабочий лист не прочитан")
        cell = self.worksheet["A1"]  # объект Cell
        cell_value = cell.value  # значение объекта Cell
        cell_row = cell.row  # строка объекта Cell
        cell_col = cell.column_letter  # столбец объекта Cell
        print(f'{cell_col}{cell_row} = {cell_value}')

    def read_and_assign_header(self):
        if not self.worksheet:
            raise AttributeError("Рабочий лист не прочитан")
        max_col = 6
        self.header_row = 1
        # назначение заголовка
        self.header = []
        #итерация по ячейкам
        for row in self.worksheet.iter_rows(min_row=self.header_row,
                                    max_row=self.header_row,
                                    min_col=1,
                                    max_col=max_col,
                                    values_only=True):
            self.header.extend(row)

    def read_and_assign_values(self):
        if not self.worksheet:
            raise AttributeError("Рабочий лист не прочитан")
        # выгрузка значений
        self.values = []
        #итерация
        for row in self.worksheet.iter_rows(min_row=self.header_row + 1,
                                    min_col=1,
                                    max_col=6,
                                    values_only=True):
            self.values.append(row)

    def output_values_as_dict(self):
        dict_vals = []
        for val in self.values:
            dict_vals.append((dict(zip(self.header, val))))
        pprint(dict_vals, indent=4)

    def output_values_as_table(self):
        table = PrettyTable(field_names=self.header)
        table.add_rows(self.values)
        print(table)


    def add_cells(self, values=[11, 'elementary', 903, 'Debian', 'apt', 2016]):
        if not self.worksheet:
            raise AttributeError("Рабочий лист не прочитан")
        if not self.workbook:
            raise AttributeError("Рабочая книга не загружена")
        self.worksheet.append(values)
        self.workbook.save(self.filename)
    
    def reread_file(self):
        self.open_file()
        self.read_sheet()
        self.read_and_assign_values()
        self.output_values_as_table()

    def find_all_based_on(self, base:str):
        """
        Дописать функцию так, чтобы можно было осуществить поиск по значению в столбце BasedOn

        Args:
            base (str): значение для фильтра
        """
        ...


    def change_ubuntu_based_on(self, base='Debian'):
        """
        Дописать функцию так, чтобы в ячейке со значением Name = Ubuntu значение в столбце BaseOn было изменено на 'Debian'

        _extended_summary_

        Args:
            base (str, optional):Defaults to 'Debian'.
        """
        ...

    def find_recently_released(self, min_year=2020):
        """
        Дописать функцию с выводом записей не ранее 2020 года
        """
        ...

    def add_new_column(self, header='Website'):
        """
        Дописать функцию для добавления столбца с данным о вебсайте

        
        Args:
            header (str, optional): наименование нового столбца
        """
        ...

    def add_new_column_data(self, websites=[
        {'EndeavourOS':'https://endeavouros.com/'}, 
        {'Manjaro':'https://manjaro.org/'},
        {'Garuda':'https://garudalinux.org/'},
        {'Debian':'http://www.debian.org/'},
        {'Fedora':'https://getfedora.org/'},
        {'Zorin':'https://www.zorinos.com/'},
        {'Mint':'https://linuxmint.com/'},
        {'MX Linux':'https://mxlinux.org/'}, 
        {'Pop!_OS':'https://system76.com/pop'},
        {'Ubuntu':'https://www.ubuntu.com/'},
        ]):
        """
        Добавить в новый столбец информацию в соответствие с записями в столбце Name

        Args:
            websites (list, optional):.
        """
        ...

    def add_color_where(self, filter:Any, color=Colors, font=FontStyles):
        """
        Добавить цвета и стили в соответствие с назначенным фильтром

        """
        for row in self.worksheet.rows:
            for cell in row:
                if cell.value:
                    ...


if __name__ == '__main__':

    app = ExcelReader('distrowatch_top_10.xlsx')
    app.open_file()
    app.read_sheet()
    app.read_and_assign_header()
    app.read_and_assign_values()
    app.output_values_as_table()
    app.add_cells()
    app.reread_file()